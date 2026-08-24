import io
import os
import uuid
from datetime import datetime, timedelta

from flask import (Flask, render_template, request, redirect, url_for, flash,
                    session, send_file, jsonify, abort)
from flask_login import (LoginManager, login_user, logout_user, login_required,
                          current_user, UserMixin)
from sqlalchemy import func, or_
from werkzeug.utils import secure_filename

from models import db, User, ScreeningRecord, Setting
from screening import (screen_one, screen_batch, match_uae_local_list,
                        load_uae_list_from_csv_text, uae_classification_en,
                        format_uae_listing_decision, uae_place_en, overall_result, risk_label,
                        load_un_list_from_text, match_un_list)
from certificate import generate_certificate_pdf_bytes

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
db_url = os.environ.get("DATABASE_URL", f"sqlite:///{os.path.join(BASE_DIR, 'screening.db')}")
if db_url.startswith("postgres://"):  # Render/Heroku-style URLs need the +psycopg2 dialect
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"


class LoginUser(UserMixin):
    def __init__(self, user):
        self.id = str(user.id)
        self.user = user


@login_manager.user_loader
def load_user(user_id):
    u = User.query.get(int(user_id))
    return LoginUser(u) if u else None


DEFAULT_THRESHOLDS = {"broad": 0.5, "balanced": 0.7, "strict": 0.85}


# ---------- bootstrap ----------

_uae_list_cache = {"records": None}


def get_uae_list():
    if _uae_list_cache["records"] is None:
        refresh_uae_list_cache()
    return _uae_list_cache["records"]


def refresh_uae_list_cache():
    csv_text = Setting.get("uae_list_csv")
    if not csv_text:
        path = os.path.join(BASE_DIR, "data", "uae_local_terrorist_list.csv")
        with open(path, encoding="utf-8-sig") as f:
            csv_text = f.read()
        Setting.set("uae_list_csv", csv_text)
        Setting.set("uae_list_updated_at", datetime.utcnow().isoformat())
        Setting.set("uae_list_filename", "uae_local_terrorist_list.csv (bundled default)")
    _uae_list_cache["records"] = load_uae_list_from_csv_text(csv_text)


_un_list_cache = {"records": None, "loaded": False}


def get_un_list():
    """Returns (records, loaded). Unlike the UAE list there's no bundled
    default - the UN's official export isn't available at a stable URL
    (see admin settings), so this starts empty until an admin uploads one."""
    if _un_list_cache["records"] is None:
        refresh_un_list_cache()
    return _un_list_cache["records"], _un_list_cache["loaded"]


def refresh_un_list_cache():
    raw = Setting.get("un_list_raw")
    is_xml = Setting.get("un_list_is_xml", "1") == "1"
    if not raw:
        _un_list_cache["records"] = []
        _un_list_cache["loaded"] = False
        return
    _un_list_cache["records"] = load_un_list_from_text(raw, is_xml)
    _un_list_cache["loaded"] = True


def get_api_key():
    return os.environ.get("OPENSANCTIONS_API_KEY", "").strip()


def init_db():
    db.create_all()
    if User.query.count() == 0:
        admin_user = os.environ.get("ADMIN_USERNAME", "admin")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "changeme123")
        u = User(username=admin_user, role="admin")
        u.set_password(admin_pass)
        db.session.add(u)
        db.session.commit()


with app.app_context():
    init_db()


# ---------- auth ----------

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        u = User.query.filter_by(username=username).first()
        if u and u.check_password(password):
            login_user(LoginUser(u), remember=True)
            return redirect(url_for("dashboard"))
        flash("Incorrect username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ---------- dashboard ----------

@app.route("/")
@login_required
def dashboard():
    total = ScreeningRecord.query.count()
    hits = ScreeningRecord.query.filter(ScreeningRecord.result_status != "Clear").count()
    pending = ScreeningRecord.query.filter_by(review_status="Pending").filter(
        ScreeningRecord.result_status != "Clear").count()
    since = datetime.utcnow() - timedelta(days=30)
    last_30d = ScreeningRecord.query.filter(ScreeningRecord.created_at >= since).count()
    recent = ScreeningRecord.query.order_by(ScreeningRecord.created_at.desc()).limit(8).all()
    uae_status = {
        "count": len(get_uae_list()),
        "updated_at": Setting.get("uae_list_updated_at", ""),
        "filename": Setting.get("uae_list_filename", ""),
    }
    un_records, un_loaded = get_un_list()
    un_status = {
        "loaded": un_loaded, "count": len(un_records),
        "updated_at": Setting.get("un_list_updated_at", ""),
        "filename": Setting.get("un_list_filename", ""),
    }
    return render_template("dashboard.html", total=total, hits=hits, pending=pending,
                            last_30d=last_30d, recent=recent, uae_status=uae_status,
                            un_status=un_status, api_key_set=bool(get_api_key()))


# ---------- single screening ----------

@app.route("/screen", methods=["GET", "POST"])
@login_required
def screen_single():
    if request.method == "GET":
        return render_template("screen.html")

    api_key = get_api_key()
    if not api_key:
        flash("No OpenSanctions API key configured on the server. Ask an admin to set OPENSANCTIONS_API_KEY.", "error")
        return redirect(url_for("screen_single"))

    name = request.form.get("name", "").strip()
    if not name:
        flash("Enter a name to screen.", "error")
        return redirect(url_for("screen_single"))

    schema = request.form.get("schema", "Person")
    country = request.form.get("country", "").strip()
    dob = request.form.get("dob", "").strip()
    threshold = float(request.form.get("threshold", 0.7))
    row = {"name": name, "schema": schema, "country": country, "birthDate": dob}

    try:
        os_matches = screen_one(row, api_key, threshold, limit=20, fetch_detail=True)
    except Exception as e:
        flash(f"Screening failed: {e}", "error")
        return redirect(url_for("screen_single"))

    uae_matches = match_uae_local_list(name, get_uae_list(), threshold)
    un_records, un_loaded = get_un_list()
    un_matches = match_un_list(name, un_records, threshold) if un_loaded else []

    ref = f"OS-SCR-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
    record = ScreeningRecord(
        reference_id=ref, subject_name=name, entity_type=schema, country=country, dob=dob,
        threshold=threshold, mode="Single",
        screened_by_id=int(current_user.id) if current_user.is_authenticated else None,
        result_status=overall_result(os_matches, uae_matches, un_matches),
    )
    record.set_matches(os_matches, uae_matches, un_matches, un_checked=un_loaded)
    db.session.add(record)
    db.session.commit()

    return redirect(url_for("record_detail", record_id=record.id))


# ---------- bulk screening ----------

@app.route("/screen/bulk", methods=["GET", "POST"])
@login_required
def screen_bulk():
    if request.method == "GET":
        return render_template("screen_bulk.html")

    api_key = get_api_key()
    if not api_key:
        flash("No OpenSanctions API key configured on the server. Ask an admin to set OPENSANCTIONS_API_KEY.", "error")
        return redirect(url_for("screen_bulk"))

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Choose a CSV or Excel file first.", "error")
        return redirect(url_for("screen_bulk"))

    threshold = float(request.form.get("threshold", 0.7))
    filename = secure_filename(file.filename)
    rows = []
    try:
        if filename.lower().endswith(".csv"):
            import csv as csv_mod
            text = file.read().decode("utf-8-sig")
            reader = csv_mod.DictReader(io.StringIO(text))
            for r in reader:
                r = {k.lower().strip(): v for k, v in r.items()}
                if r.get("name", "").strip():
                    rows.append({
                        "name": r["name"].strip(),
                        "schema": "Company" if r.get("schema", "").strip().lower() in
                                  ("company", "entity", "organization", "org") else "Person",
                        "country": r.get("country", "").strip(),
                        "birthDate": r.get("birthdate", "").strip(),
                        "filename": r.get("filename", "").strip(),
                    })
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for r in ws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(header, [("" if v is None else str(v)) for v in r]))
                if d.get("name", "").strip():
                    rows.append({
                        "name": d["name"].strip(),
                        "schema": "Company" if d.get("schema", "").strip().lower() in
                                  ("company", "entity", "organization", "org") else "Person",
                        "country": d.get("country", "").strip(),
                        "birthDate": d.get("birthdate", "").strip(),
                        "filename": d.get("filename", "").strip(),
                    })
    except Exception as e:
        flash(f"Could not read file: {e}", "error")
        return redirect(url_for("screen_bulk"))

    if not rows:
        flash("No rows with a 'name' value found in that file.", "error")
        return redirect(url_for("screen_bulk"))
    if len(rows) > 500:
        flash(f"File has {len(rows)} rows; please split into batches of 500 or fewer.", "error")
        return redirect(url_for("screen_bulk"))

    try:
        os_results = screen_batch(rows, api_key, threshold, limit=20, fetch_detail=True)
    except Exception as e:
        flash(f"Screening failed: {e}", "error")
        return redirect(url_for("screen_bulk"))

    batch_id = f"BATCH-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
    uae_list = get_uae_list()
    un_records, un_loaded = get_un_list()
    created = 0
    for i, row in enumerate(rows):
        os_matches = os_results.get(i, [])
        uae_matches = match_uae_local_list(row["name"], uae_list, threshold)
        un_matches = match_un_list(row["name"], un_records, threshold) if un_loaded else []
        ref = f"{batch_id}-{i+1:04d}"
        record = ScreeningRecord(
            reference_id=ref, batch_id=batch_id, subject_name=row["name"],
            entity_type=row["schema"], country=row["country"], dob=row["birthDate"],
            threshold=threshold, mode="Bulk",
            screened_by_id=int(current_user.id) if current_user.is_authenticated else None,
            result_status=overall_result(os_matches, uae_matches, un_matches),
        )
        record.set_matches(os_matches, uae_matches, un_matches, un_checked=un_loaded)
        db.session.add(record)
        created += 1
    db.session.commit()

    flash(f"Screened {created} names. View results below.", "success")
    return redirect(url_for("results", batch=batch_id))


# ---------- results / search ----------

@app.route("/results")
@login_required
def results():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    review = request.args.get("review", "").strip()
    mode = request.args.get("mode", "").strip()
    batch = request.args.get("batch", "").strip()
    page = max(int(request.args.get("page", 1)), 1)
    per_page = 25

    query = ScreeningRecord.query
    if q:
        query = query.filter(or_(ScreeningRecord.subject_name.ilike(f"%{q}%"),
                                  ScreeningRecord.reference_id.ilike(f"%{q}%")))
    if status:
        query = query.filter(ScreeningRecord.result_status == status)
    if review:
        query = query.filter(ScreeningRecord.review_status == review)
    if mode:
        query = query.filter(ScreeningRecord.mode == mode)
    if batch:
        query = query.filter(ScreeningRecord.batch_id == batch)

    total = query.count()
    records = (query.order_by(ScreeningRecord.created_at.desc())
               .offset((page - 1) * per_page).limit(per_page).all())
    total_pages = max((total + per_page - 1) // per_page, 1)

    return render_template("results.html", records=records, q=q, status=status, review=review,
                            mode=mode, batch=batch, page=page, total_pages=total_pages, total=total)


@app.route("/results/export.csv")
@login_required
def export_results_csv():
    import csv as csv_mod
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    batch = request.args.get("batch", "").strip()
    query = ScreeningRecord.query
    if q:
        query = query.filter(or_(ScreeningRecord.subject_name.ilike(f"%{q}%"),
                                  ScreeningRecord.reference_id.ilike(f"%{q}%")))
    if status:
        query = query.filter(ScreeningRecord.result_status == status)
    if batch:
        query = query.filter(ScreeningRecord.batch_id == batch)
    records = query.order_by(ScreeningRecord.created_at.desc()).all()

    buf = io.StringIO()
    writer = csv_mod.writer(buf)
    writer.writerow(["Reference ID", "Subject", "Type", "Country", "Mode", "Result",
                      "Review Status", "Match Count", "Screened By", "Date"])
    for r in records:
        writer.writerow([r.reference_id, r.subject_name, r.entity_type, r.country, r.mode,
                          r.result_status, r.review_status, r.match_count,
                          r.screened_by.username if r.screened_by else "", r.created_at])
    mem = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                      download_name=f"screening_results_{datetime.utcnow().strftime('%Y%m%d')}.csv")


@app.route("/results/<int:record_id>")
@login_required
def record_detail(record_id):
    record = ScreeningRecord.query.get_or_404(record_id)
    os_matches = sorted(record.os_matches(), key=lambda m: m.get("score", 0), reverse=True)
    uae_matches = sorted(record.uae_matches(), key=lambda m: m["score"], reverse=True)
    un_matches = sorted(record.un_matches(), key=lambda m: m["score"], reverse=True)
    for m in uae_matches:
        m["record"]["classification_en"] = uae_classification_en(m["record"].get("classification_ar", ""))
        m["record"]["listing_decision_en"] = format_uae_listing_decision(m["record"].get("listing_decision", ""))
    return render_template("record_detail.html", record=record, os_matches=os_matches,
                            uae_matches=uae_matches, un_matches=un_matches, risk_label=risk_label)


@app.route("/results/<int:record_id>/review", methods=["POST"])
@login_required
def update_review(record_id):
    record = ScreeningRecord.query.get_or_404(record_id)
    record.review_status = request.form.get("review_status", "Pending")
    record.review_notes = request.form.get("review_notes", "").strip()
    record.reviewed_by_id = int(current_user.id)
    record.reviewed_at = datetime.utcnow()
    db.session.commit()
    flash("Review saved.", "success")
    return redirect(url_for("record_detail", record_id=record_id))


@app.route("/results/<int:record_id>/pdf")
@login_required
def download_pdf(record_id):
    record = ScreeningRecord.query.get_or_404(record_id)
    pdf_bytes = generate_certificate_pdf_bytes(record)
    filename = f"{record.subject_name.replace(' ', '_')}_{record.reference_id}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                      as_attachment=True, download_name=filename)


# ---------- admin ----------

def require_admin():
    u = User.query.get(int(current_user.id))
    if not u or not u.is_admin:
        abort(403)


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
def admin_settings():
    require_admin()
    if request.method == "POST":
        uae_file = request.files.get("uae_list_file")
        if uae_file and uae_file.filename:
            text = uae_file.read().decode("utf-8-sig")
            Setting.set("uae_list_csv", text)
            Setting.set("uae_list_updated_at", datetime.utcnow().isoformat())
            Setting.set("uae_list_filename", secure_filename(uae_file.filename))
            _uae_list_cache["records"] = None
            flash("UAE Local Terrorist List updated.", "success")

        un_file = request.files.get("un_list_file")
        if un_file and un_file.filename:
            fname = secure_filename(un_file.filename)
            is_xml = fname.lower().endswith(".xml")
            text = un_file.read().decode("utf-8", errors="ignore")
            test_records = load_un_list_from_text(text, is_xml)
            if not test_records:
                flash("Could not parse any records from that UN list file \u2014 check it's the "
                      "official XML or HTML export from scsanctions.un.org/consolidated/. "
                      "Nothing was changed.", "error")
            else:
                Setting.set("un_list_raw", text)
                Setting.set("un_list_is_xml", "1" if is_xml else "0")
                Setting.set("un_list_updated_at", datetime.utcnow().isoformat())
                Setting.set("un_list_filename", fname)
                _un_list_cache["records"] = None
                flash(f"UN Consolidated List updated \u2014 {len(test_records)} records loaded.", "success")

        return redirect(url_for("admin_settings"))

    uae_status = {
        "count": len(get_uae_list()),
        "updated_at": Setting.get("uae_list_updated_at", ""),
        "filename": Setting.get("uae_list_filename", ""),
    }
    un_records, un_loaded = get_un_list()
    un_status = {
        "loaded": un_loaded, "count": len(un_records),
        "updated_at": Setting.get("un_list_updated_at", ""),
        "filename": Setting.get("un_list_filename", ""),
    }
    users = User.query.order_by(User.username).all()
    return render_template("settings.html", uae_status=uae_status, un_status=un_status,
                            users=users, api_key_set=bool(get_api_key()))


@app.route("/admin/users", methods=["POST"])
@login_required
def create_user():
    require_admin()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "analyst")
    if not username or not password:
        flash("Username and password are required.", "error")
        return redirect(url_for("admin_settings"))
    if User.query.filter_by(username=username).first():
        flash("That username is already taken.", "error")
        return redirect(url_for("admin_settings"))
    u = User(username=username, role=role)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    flash(f"User '{username}' created.", "success")
    return redirect(url_for("admin_settings"))


if __name__ == "__main__":
    app.run(debug=True, port=5000)
